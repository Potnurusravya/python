from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from tkinter.ttk import Combobox
import pathlib
import openpyxl  # pip install openpyxl
from openpyxl import Workbook
import os
from PIL import ImageTk,Image  # pip install pillow

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"
root=Tk()
root.title("Student Registration System")
root.geometry("1250x720+210+100")
root.config(bg=background)


#program for student registration

file=pathlib.Path("Student_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No"
    sheet['B1']="Name"
    sheet['C1']="Study"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Religion"
    sheet['H1']="Course"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"
    file.save('Student_data.xlsx')


def selection():
    global gender
    value=radio.get()
    if value==1:
            gender="Male"
        
    elif value==2:
        gender="Female"
    else:
        gender="Others"
    print(gender)

#functions for btns

#exit window
def Exit():
    root.destroy()

#show_profile_img
def show_img():
    global filename
    global profile_pic
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select Image",filetypes=(("JPG File","*.jpg"),("PNG File","*.png")))
    profile_pic=(Image.open(filename))
    resize_img=profile_pic.resize((190,190))
    Photo2=ImageTk.PhotoImage(resize_img)
    lbl.config(image=Photo2)
    lbl.image=Photo2
    print(filename)
    print(profile_pic)

#Register No
def register_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value
    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set('1')

#all clear
def clear():
    Name.set('')
    DOB.set('')
    Religion.set('')
    Course.set('')
    F_Name.set('')
    F_Occupation.set('')
    M_Name.set('')
    M_Occupation.set('')
    study.set("Select Qualification")
    register_no()
    save_btn.config(state='normal')   
    img1=PhotoImage(file='profile.png')
    lbl.config(image=img1)
    lbl.image=img1
    profile_pic=''
    
def save_data():
    R1=Registration.get()
    N1=Name.get()
    S1=study.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender")
    D2=DOB.get()
    D1=Date.get()
    Rel=Religion.get()
    C1=Course.get()
    father_name=F_Name.get()
    mother_name=M_Name.get()
    F1=F_Occupation.get()
    M1=M_Occupation.get()
    if N1=="" or S1=="Select Qualification" or D2=="" or Rel=="" or C1=="" or father_name=="" or mother_name=="" or F1=="" or M1=="":
        messagebox.showerror("Error","Please Provide All Details !!")
    else:
        file=openpyxl.load_workbook("Student_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)    
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=S1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)
        sheet.cell(column=7,row=sheet.max_row,value=Rel)
        sheet.cell(column=8,row=sheet.max_row,value=C1)
        sheet.cell(column=9,row=sheet.max_row,value=father_name)
        sheet.cell(column=10,row=sheet.max_row,value=mother_name)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)
        file.save('Student_data.xlsx')
        try:
            profile_pic.save("studentsphotos/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("Info","Profile picture is not available !!!")
        messagebox.showinfo("Info","Successfully data Saved !!")
        clear() # After Save the data clear the entry box and image
        register_no() # this is for it will recheck and reissue new no.
        
### Search for data         
def search():
    text=Search.get()
    clear() # to clear all the data already available in entry box and others
    save_btn.config(state='disabled') 
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    
    for row in sheet.rows:
        # print(row)
        if row[0].value==int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
            # print(reg_no_position,reg_number)
    try:
        x1=sheet.cell(row=int(reg_number),column=1).value
        x2=sheet.cell(row=int(reg_number),column=2).value
        x3=sheet.cell(row=int(reg_number),column=3).value
        x4=sheet.cell(row=int(reg_number),column=4).value
        x5=sheet.cell(row=int(reg_number),column=5).value
        x6=sheet.cell(row=int(reg_number),column=6).value
        x7=sheet.cell(row=int(reg_number),column=7).value
        x8=sheet.cell(row=int(reg_number),column=8).value
        x9=sheet.cell(row=int(reg_number),column=9).value
        x10=sheet.cell(row=int(reg_number),column=10).value
        x11=sheet.cell(row=int(reg_number),column=11).value
        x12=sheet.cell(row=int(reg_number),column=12).value
        Registration.set(x1)
        Name.set(x2)
        study.set(x3)
        if x4=='Female':
            R2.select()
        elif x4=='Male':
            R1.select()
        else:
            R3.select()
        DOB.set(x5)
        Date.set(x6)
        Religion.set(x7)
        Course.set(x8)
        F_Name.set(x9)
        M_Name.set(x10)
        F_Occupation.set(x11)
        M_Occupation.set(x12)
        try:
            img=Image.open("studentsphotos/"+str(x1)+".jpg")  # we done this to take image name ,same as Registration no.
            resize_img=img.resize((190,190))
            Photo2=ImageTk.PhotoImage(resize_img)
            lbl.config(image=Photo2)
            lbl.image=Photo2
        except:
            pass
    except:
            messagebox.showerror("Invalid","Invalid Registeration No")
    
## for update Data
def update_data():
    R1=Registration.get()
    N1=Name.get()
    S1=study.get()
    selection()
    G1=gender
    D2=DOB.get()
    D1=Date.get()
    Rel=Religion.get()
    C1=Course.get()
    father_name=F_Name.get()
    mother_name=M_Name.get()
    F1=F_Occupation.get()
    M1=M_Occupation.get()   
    
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    
    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            # print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
            
    sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=S1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=Rel)
    sheet.cell(column=8,row=int(reg_number),value=C1)
    sheet.cell(column=9,row=int(reg_number),value=father_name)
    sheet.cell(column=10,row=int(reg_number),value=mother_name)
    sheet.cell(column=11,row=int(reg_number),value=F1)
    sheet.cell(column=12,row=int(reg_number),value=M1)
    file.save("Student_data.xlsx")
    try:
        profile_pic.save("studentsphotos/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("Update","Update Successfully !!")
    clear() # after that clear all the boxes

#Interface for student registration

#top frame
Label(root,text="Email: indhuvakada@gmail.com",width=10,height=2,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464", fg="#fff" ,font=("arial",20,"bold")).pack(side=TOP,fill=X)



#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font=("arial",20)).place(x=810,y=50)
imgicon3=PhotoImage(file="Project/search.png")
Srch=Button(root,compound=LEFT,image=imgicon3,width=50,height=33,bd=2,bg="#68ddfa",font=("arial",13,"bold"),command=search)
Srch.place(x=1042,y=50)

#update image
imgicon4=PhotoImage(file="Project/refresh1.png")
Update_btn=Button(root,text="Update Data",compound=LEFT,font=("arial",10,"bold"),image=imgicon4,bg="lightblue",command=update_data)
Update_btn.place(x=50,y=47)

#date and registration number
Label(root,text='Registration No:',font=('arial',13),fg=framebg,bg=background,).place(x=50,y=140)
Label(root,text='Date:',font=('arial',13),fg=framebg,bg=background).place(x=500,y=140)
Registration=IntVar()
Date=StringVar()
reg_entry=Entry(root,textvariable=Registration,width=15,font=("arial",10),state='readonly')
reg_entry.place(x=175,y=140)
register_no()



today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font=("arial",10),state='readonly')
date_entry.place(x=550,y=140)
Date.set(d1)

#Student Details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Enter FullName:",font=('arial',13),fg=framefg,bg=framebg).place(x=20,y=30)
Label(obj,text="Date of Birth:",font=('arial',13),fg=framefg,bg=framebg).place(x=20,y=90)
Label(obj,text="Gender:",font=('arial',13),fg=framefg,bg=framebg).place(x=20,y=150)

Label(obj,text="Qualification:",font=('arial',13),fg=framefg,bg=framebg).place(x=500,y=30)
Label(obj,text="Religion:",font=('arial',13),fg=framefg,bg=framebg).place(x=500,y=90)
Label(obj,text="Course:",font=('arial',13),fg=framefg,bg=framebg).place(x=500,y=150)

# Student Entries
Name=StringVar()
name_entry=Entry(obj,textvariable=Name,font=("arial",10),width=20,)
name_entry.place(x=150,y=32)
DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,font=("arial",10),width=20,)
dob_entry.place(x=150,y=92)

radio=IntVar()
R1=Radiobutton(obj,text="Male",variable=radio,value=1,fg=framefg,bg=framebg,command=selection)
R1.place(x=150,y=152)
R2=Radiobutton(obj,text="Female",variable=radio,value=2,fg=framefg,bg=framebg,command=selection)
R2.place(x=200,y=152)
R3=Radiobutton(obj,text="Others",variable=radio,value=3,fg=framefg,bg=framebg,command=selection)
R3.place(x=265,y=152)

data=["Intermediate","Diploma","Bachelor's degree","Engineering","MCA","MBA","Phd"]
study=Combobox(obj,values=data,font=("roboto",10),width=17,state='r')
study.place(x=600,y=31)
study.set("Select Qualification")

Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,font=("arial",10),width=20,)
religion_entry.place(x=600,y=92)

Course=StringVar()
course_entry=Entry(obj,textvariable=Course,font=("arial",10),width=20,)
course_entry.place(x=600,y=152)


#Parent's Details
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=230,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name:",font=('arial',13),fg=framefg,bg=framebg).place(x=20,y=30)
Label(obj2,text="Father's Occupation:",font=('arial',13),fg=framefg,bg=framebg).place(x=20,y=90)

Label(obj2,text="Mother's Name:",font=('arial',13),fg=framefg,bg=framebg).place(x=500,y=30)
Label(obj2,text="Mother's Occupation:",font=('arial',13),fg=framefg,bg=framebg).place(x=500,y=90)

# Parent Entries
F_Name=StringVar()
f_name_entry=Entry(obj2,textvariable=F_Name,width=20,font=("arial",10))
f_name_entry.place(x=160,y=32)

F_Occupation=StringVar()
f_occupation_entry=Entry(obj2,textvariable=F_Occupation,width=20,font=("arial",10))
f_occupation_entry.place(x=180,y=92)

M_Name=StringVar()
m_name_entry=Entry(obj2,textvariable=M_Name,width=20,font=("arial",10))
m_name_entry.place(x=630,y=32)

M_Occupation=StringVar()
m_occupation_entry=Entry(obj2,textvariable=M_Occupation,width=20,font=("arial",10))
m_occupation_entry.place(x=670,y=92)

#Photo Upload
#profile image
f=Frame(root,bd=3,bg='black',width=200,height=195,relief=GROOVE)
f.place(x=1000,y=150)

profile_img=PhotoImage(file='Project/profile.png')
lbl=Label(f,bg="black",image=profile_img)
lbl.place(x=0,y=0)

#Buttons
Button(root,text="Upload",width=19,height=2,font=("arial",12,"bold"),bg="lightblue",command=show_img).place(x=1000,y=370)
save_btn=Button(root,text="Save",width=19,height=2,font=("arial",12,"bold"),bg="lightgreen",command=save_data)
save_btn.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font=("arial",12,"bold"),bg="lightpink",command=clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font=("arial",12,"bold"),bg="red",fg='white',command=Exit).place(x=1000,y=610)
root.mainloop()